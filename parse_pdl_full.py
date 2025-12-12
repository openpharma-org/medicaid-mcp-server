
import openpyxl
import json

wb = openpyxl.load_workbook('il_pdl_latest.xlsx', data_only=True)
sheet = wb.active

drugs = []
for row in sheet.iter_rows(min_row=41, max_row=sheet.max_row):
    values = [cell.value for cell in row]
    if not values or not values[1]:
        continue

    drug_name = str(values[1] or '').strip()
    dosage_form = str(values[2] or '').strip()
    pdl_status = str(values[4] or values[5] or '').strip()

    if drug_name:
        drugs.append({
            'drug_name': drug_name,
            'dosage_form': dosage_form,
            'pdl_status': pdl_status
        })

print(json.dumps(drugs))
wb.close()
