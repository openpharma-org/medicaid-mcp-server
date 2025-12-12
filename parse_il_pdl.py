
import openpyxl
import json
import sys

try:
    wb = openpyxl.load_workbook('il_pdl_latest.xlsx', data_only=True)
    sheet = wb.active

    # Headers at row 40
    drugs = []
    for row in sheet.iter_rows(min_row=41, max_row=sheet.max_row):
        values = [cell.value for cell in row]
        if not values or not values[1]:  # Skip empty rows
            continue

        drug_class = str(values[0] or '').strip()
        drug_name = str(values[1] or '').strip()
        dosage_form = str(values[2] or '').strip()
        pdl_status_1 = str(values[3] or '').strip()
        pdl_status_2 = str(values[4] or '').strip()
        pdl_status_3 = str(values[5] or '').strip()

        # Combine PDL status fields
        pdl_status = pdl_status_1 or pdl_status_2 or pdl_status_3

        if drug_name:
            drugs.append({
                'drug_class': drug_class,
                'drug_name': drug_name,
                'dosage_form': dosage_form,
                'pdl_status': pdl_status
            })

    print(json.dumps({'success': True, 'drugs': drugs, 'count': len(drugs)}))
    wb.close()

except Exception as e:
    print(json.dumps({'success': False, 'error': str(e)}), file=sys.stderr)
    sys.exit(1)
