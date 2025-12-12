#!/usr/bin/env python3
"""
Inspect Illinois Medicaid PDL Excel file structure
"""

import openpyxl
import sys

def inspect_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename, data_only=False)

        print("=" * 80)
        print(f"Illinois Medicaid PDL: {filename}")
        print("=" * 80)

        print(f"\nWorkbook sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            print(f"Dimensions: {sheet.dimensions}")
            print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

            # Try to find headers (check rows 1, 38-42)
            print("\nChecking for headers:")
            for i in [1, 38, 39, 40, 41, 42]:
                if i <= sheet.max_row:
                    row_values = [cell.value for cell in sheet[i]]
                    non_empty = [v for v in row_values if v is not None]
                    if non_empty:
                        print(f"Row {i}: {row_values[:10]}")

            # Sample data (rows around 40)
            print("\nSample data (rows 40-45):")
            for i in range(40, min(46, sheet.max_row + 1)):
                row_values = [cell.value for cell in sheet[i]]
                non_empty = [v for v in row_values if v is not None]
                if non_empty:
                    print(f"Row {i}: {row_values}")

            # Count non-empty rows
            non_empty_count = 0
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                if any(cell.value is not None for cell in row):
                    non_empty_count += 1
            print(f"\nTotal non-empty rows: {non_empty_count}")

        wb.close()

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    inspect_excel("il_pdl_sample.xlsx")
